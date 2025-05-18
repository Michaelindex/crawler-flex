"""
Exportador de dados para formato Excel.
Responsável por exportar os dados coletados para arquivos Excel.
"""

import logging
import os
from datetime import datetime
from typing import Dict, Any, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

class ExcelExporter:
    """
    Exportador de dados para formato Excel.
    """
    
    def __init__(self):
        """Inicializa o exportador Excel."""
        self.output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'data', 'output')
        
        # Garantir que o diretório de saída existe
        os.makedirs(self.output_dir, exist_ok=True)
    
    def export(self, data: List[Dict[str, Any]], filename: str = None) -> str:
        """
        Exporta os dados para um arquivo Excel.
        
        Args:
            data: Lista de dicionários com os dados a serem exportados
            filename: Nome do arquivo (opcional)
            
        Returns:
            Caminho do arquivo exportado
        """
        if not data:
            logger.warning("Nenhum dado para exportar")
            return ""
        
        # Gerar nome de arquivo se não fornecido
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"empresas_{timestamp}.xlsx"
        
        # Caminho completo do arquivo
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Converter dados para DataFrame
            df = pd.DataFrame(data)
            
            # Exportar com formatação
            self._export_with_formatting(df, filepath)
            return filepath
        
        except Exception as e:
            logger.error(f"Erro ao exportar dados para Excel: {e}")
            return ""
    
    def _export_with_formatting(self, df: pd.DataFrame, filepath: str) -> None:
        """
        Exporta os dados para Excel com formatação avançada.
        
        Args:
            df: DataFrame com os dados
            filepath: Caminho do arquivo de saída
        """
        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas"
        
        # Adicionar os dados ao worksheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Definir estilos
        try:
            # Estilo para cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Estilo para bordas
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Aplicar estilos ao cabeçalho
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Aplicar bordas e alinhamento às células de dados
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column_letter].width = adjusted_width
        
        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação: {e}")
        
        # Salvar o arquivo
        wb.save(filepath)
        logger.info(f"Dados exportados com formatação para {filepath}")
    
    def export_csv(self, data: List[Dict[str, Any]], filename: str = None) -> str:
        """
        Exporta os dados para um arquivo CSV.
        
        Args:
            data: Lista de dicionários com os dados a serem exportados
            filename: Nome do arquivo (opcional)
            
        Returns:
            Caminho do arquivo exportado
        """
        if not data:
            logger.warning("Nenhum dado para exportar")
            return ""
        
        # Gerar nome de arquivo se não fornecido
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"empresas_{timestamp}.csv"
        
        # Caminho completo do arquivo
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            # Converter dados para DataFrame
            df = pd.DataFrame(data)
            
            # Exportar para CSV
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
            logger.info(f"Dados exportados para CSV: {filepath}")
            
            return filepath
        
        except Exception as e:
            logger.error(f"Erro ao exportar dados para CSV: {e}")
            return ""
